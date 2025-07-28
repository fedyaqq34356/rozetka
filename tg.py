import argparse
import os
import re
import sys
import time
from datetime import datetime

try:
    import cloudscraper
except ImportError:
    print("[ПОМИЛКА] Потрібно встановити cloudscraper: pip install cloudscraper")
    raise



try:
    import openpyxl
    from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("[ПОМИЛКА] Потрібно встановити openpyxl: pip install openpyxl")
    raise

try:
    from bs4 import BeautifulSoup
    _HAVE_BS4 = True
except ImportError:
    _HAVE_BS4 = False


class RozetkaStockChecker:
    def __init__(self, debug=False, delay=0.7):
        self.scraper = cloudscraper.create_scraper(
            browser={
                'browser': 'chrome',
                'platform': 'windows',
                'mobile': False,
                'desktop': True
            },
            delay=10,
            interpreter='js2py'
        )
        self.base_headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'ru,en;q=0.9,en-GB;q=0.8,en-US;q=0.7,uk;q=0.6',
            'Content-Type': 'application/json',
            'Origin': 'https://rozetka.com.ua',
            'Referer': 'https://rozetka.com.ua/',
            'X-Requested-With': 'XMLHttpRequest',
        }
        self.debug = debug
        self.delay = delay
        self.reset_session_state()

    def reset_session_state(self):
        """Очищаємо стан сесії перед перевіркою нового товару"""
        self.csrf_token = None
        self.purchase_id = None
        # Очищаємо кукі та створюємо новий scraper для кожного товару
        self.scraper = cloudscraper.create_scraper()

    def get_csrf_token(self):
        try:
            # Make request with additional headers to mimic a browser
            headers = self.base_headers.copy()
            headers.update({
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'same-origin',
                'Sec-Fetch-Dest': 'document',
                'Upgrade-Insecure-Requests': '1'
            })
            resp = self.scraper.get('https://rozetka.com.ua/', headers=headers, timeout=10)
            resp.raise_for_status()

            # Check cookies
            cookies = self.scraper.cookies.get_dict()
            if self.debug:
                print("[ДЕБАГ] Все куки:", cookies)

            possible_csrf_names = ['_uss-csrf', 'csrf-token', 'X-CSRF-TOKEN', 'csrf_token', '_token']
            for csrf_name in possible_csrf_names:
                if csrf_name in cookies:
                    self.csrf_token = cookies[csrf_name]
                    if self.debug:
                        print(f"[ДЕБАГ] Найден CSRF токен '{csrf_name}': {self.csrf_token}")
                    return True

            # Parse HTML for token
            html = resp.text
            csrf_patterns = [
                r'name="csrf-token"\s+content="([^"]+)"',
                r'"csrf_token"\s*:\s*"([^"]+)"',
                r'_uss-csrf["\']?\s*[:=]\s*["\']([^"\']+)',
                r'csrfToken["\']?\s*[:=]\s*["\']([^"\']+)',
                r'meta\[name=["\']?_?csrf[-_]?token["\']?\]\s*content=["\']([^"\']+)["\']'
            ]
            for pattern in csrf_patterns:
                match = re.search(pattern, html, re.I)
                if match:
                    self.csrf_token = match.group(1)
                    if self.debug:
                        print(f"[ДЕБАГ] CSRF токен найден в HTML: {self.csrf_token}")
                    return True

            # Try API request to set cookies
            test_url = 'https://uss.rozetka.com.ua/session/cart-se/clear?country=UA&lang=ua'
            test_resp = self.scraper.post(test_url, json={}, headers=self.base_headers, timeout=10)
            cookies = self.scraper.cookies.get_dict()
            for csrf_name in possible_csrf_names:
                if csrf_name in cookies:
                    self.csrf_token = cookies[csrf_name]
                    if self.debug:
                        print(f"[ДЕБАГ] CSRF токен получен после тестового запроса: {self.csrf_token}")
                    return True

            if self.debug:
                print("[ДЕБАГ] CSRF токен не найден")
            return False

        except Exception as e:
            if self.debug:
                print(f"[CSRF] Помилка: {e}")
            return False
            
            # Если ничего не нашли, попробуем сделать запрос к API без токена
            # Иногда первый запрос может установить нужные куки
            try:
                test_url = 'https://uss.rozetka.com.ua/session/cart-se/clear?country=UA&lang=ua'
                test_resp = self.scraper.post(test_url, json={})
                
                # Проверяем куки еще раз
                cookies = self.scraper.cookies.get_dict()
                for csrf_name in possible_csrf_names:
                    if csrf_name in cookies:
                        self.csrf_token = cookies[csrf_name]
                        if self.debug:
                            print(f"[ДЕБАГ] CSRF токен получен после тестового запроса: {self.csrf_token}")
                        return True
            except:
                pass
            
            if self.debug:
                print("[ДЕБАГ] CSRF токен не найден")
            return False
            
        except Exception as e:
            print(f"[CSRF] Помилка: {e}")
            return False

    @staticmethod
    def extract_product_id(url: str):
        match = re.search(r'/p(\d+)/', url)
        return int(match.group(1)) if match else None

    def _ensure_csrf(self):
        if not self.csrf_token:
            if not self.get_csrf_token():
                raise RuntimeError("Не вдалось отримати CSRF токен (_uss-csrf)")

    def clear_cart(self):
        """Очищаємо корзину перед додаванням нового товару"""
        try:
            if not self.csrf_token:
                return
            
            url = 'https://uss.rozetka.com.ua/session/cart-se/clear?country=UA&lang=ua'
            headers = self.base_headers.copy()
            headers['CSRF-Token'] = self.csrf_token
            
            r = self.scraper.post(url, json={}, headers=headers)
            if self.debug:
                print("[ДЕБАГ] clear_cart статус:", r.status_code)
                print("[ДЕБАГ] clear_cart тіло:", r.text[:300])
        except Exception as e:
            if self.debug:
                print(f"[clear_cart] Помилка: {e}")

    def add_to_cart(self, product_id):
        self._ensure_csrf()
        
        # Очищаємо корзину перед додаванням
        self.clear_cart()
        
        url = 'https://uss.rozetka.com.ua/session/cart-se/add?country=UA&lang=ua'
        headers = self.base_headers.copy()
        headers['CSRF-Token'] = self.csrf_token
        payload = [{"goods_id": product_id, "quantity": 1}]
        
        try:
            r = self.scraper.post(url, json=payload, headers=headers)
            if self.debug:
                print(f"[ДЕБАГ] add_to_cart статус для товару {product_id}:", r.status_code)
                print("[ДЕБАГ] add_to_cart тіло:", r.text[:500])
            
            if r.status_code == 200:
                data = r.json()
                goods_items = data.get('purchases', {}).get('goods')
                if goods_items and len(goods_items) > 0:
                    # ВИПРАВЛЕННЯ: Перевіряємо що товар той самий
                    for item in goods_items:
                        if item.get('goods', {}).get('id') == product_id:
                            self.purchase_id = item['id']
                            if self.debug:
                                print(f"[ДЕБАГ] purchase_id встановлено: {self.purchase_id}")
                            return data
                    
                    print(f"[ПОПЕРЕДЖЕННЯ] Товар {product_id} не знайдено в корзині")
                    return None
                else:
                    print("[ПОПЕРЕДЖЕННЯ] Порожня корзина після додавання")
                    return None
            return None
        except Exception as e:
            print(f"[add_to_cart] Помилка для товару {product_id}: {e}")
            return None

    def update_quantity(self, quantity):
        if not self.purchase_id or not self.csrf_token:
            if self.debug:
                print(f"[update_quantity] Відсутні дані: purchase_id={self.purchase_id}, csrf_token={bool(self.csrf_token)}")
            return None
            
        url = 'https://uss.rozetka.com.ua/session/cart-se/edit-quantity?country=UA&lang=ua'
        headers = self.base_headers.copy()
        headers['CSRF-Token'] = self.csrf_token
        payload = [{"purchase_id": self.purchase_id, "quantity": quantity}]
        
        try:
            r = self.scraper.post(url, json=payload, headers=headers)
            if self.debug:
                print(f"[ДЕБАГ] update_quantity({quantity}) статус:", r.status_code)
                print("[ДЕБАГ] update_quantity тіло:", r.text[:500])
            if r.status_code == 200:
                return r.json()
            return None
        except Exception as e:
            print(f"[update_quantity] Помилка: {e}")
            return None

    def binary_search_max_stock(self, product_id, max_attempts=100, upper_bound=10000):
        if self.debug:
            print(f"[БП] Починаємо бінарний пошук для товару {product_id}")
        
        add_data = self.add_to_cart(product_id)
        if not add_data:
            print(f"[БП] Не вдалося додати товар {product_id} до корзини")
            return None, None

        left, right = 1, upper_bound
        max_available = 0

        for attempt in range(max_attempts):
            if left > right:
                break
                
            mid = (left + right) // 2
            if self.debug:
                print(f"[БП] #{attempt+1} товар {product_id} -> тестуємо кількість {mid}")
                
            data = self.update_quantity(mid)
            if not data:
                if self.debug:
                    print(f"[БП] Не отримано відповіді на {mid}")
                break
                
            time.sleep(self.delay)
            
            errors = data.get('error_messages') or []
            not_enough = False
            
            for err in errors:
                if self.debug:
                    print(f"[БП] Помилка: {err}")
                if err.get('code') == 3002:  # Код помилки "недостатньо товару"
                    not_enough = True
                    break
                    
            if not_enough:
                right = mid - 1
                if self.debug:
                    print(f"[БП] Недостатньо товару на {mid}, зменшуємо праву межу до {right}")
            else:
                max_available = mid
                left = mid + 1
                if self.debug:
                    print(f"[БП] {mid} товарів доступно, збільшуємо ліву межу до {left}")

        if self.debug:
            print(f"[БП] Результат для товару {product_id}: {max_available}")
            
        return max_available, add_data

    def parse_category_from_html(self, product_url, category_id):
        """Покращена функція парсингу категорії з більш точними селекторами"""
        try:
            resp = self.scraper.get(product_url, timeout=15)
            html = resp.text
            
            if self.debug:
                print(f"[parse_category] Шукаємо категорію ID: {category_id}")
            
            # Спочатку шукаємо в JSON-LD структурованих даних
            json_patterns = [
                rf'"name"\s*:\s*"([^"]*)"[^}}]*"@id"\s*:\s*"[^"]*/{category_id}/',
                rf'"@id"\s*:\s*"[^"]*/{category_id}/[^"]*"[^}}]*"name"\s*:\s*"([^"]*)"',
                rf'"categoryId"\s*:\s*{category_id}[^}}]*"name"\s*:\s*"([^"]*)"',
                rf'"id"\s*:\s*{category_id}[^}}]*"title"\s*:\s*"([^"]*)"'
            ]
            
            for pattern in json_patterns:
                matches = re.finditer(pattern, html, re.I | re.S)
                for match in matches:
                    category_name = match.group(1).strip()
                    if category_name and len(category_name) > 2 and len(category_name) < 100:
                        if self.debug:
                            print(f"[parse_category] Знайдено в JSON: '{category_name}'")
                        return category_name
            
            if _HAVE_BS4:
                soup = BeautifulSoup(html, 'html.parser')
                
                # Шукаємо у breadcrumbs (найбільш надійно)
                breadcrumb_selectors = [
                    '.breadcrumbs a',
                    '.rz-breadcrumbs a',
                    '[data-testid="breadcrumbs"] a',
                    '.catalog-heading a',
                    '.breadcrumb a',
                    'nav a',
                    '.rz-catalog-breadcrumbs a',
                    'a[rzrelnofollow].black-link',  # Добавлен селектор для Angular-компонентов
                    'a[href*="/c"]'
                ]
                
                for selector in breadcrumb_selectors:
                    try:
                        elements = soup.select(selector)
                        for element in elements:
                            href = element.get('href', '')
                            # Шукаємо точний збіг category_id в URL
                            if re.search(rf'/c{category_id}(?:/|$)', href):
                                text = element.get_text(strip=True)
                                if text and len(text) > 2 and len(text) < 100:
                                    # Фільтруємо технічні елементи
                                    if not any(skip in text.lower() for skip in ['>', '<', 'img', 'svg', 'icon', 'span']):
                                        if self.debug:
                                            print(f"[parse_category] Знайдено в breadcrumbs: '{text}'")
                                        return text
                    except Exception as e:
                        if self.debug:
                            print(f"[parse_category] Помилка breadcrumb селектора {selector}: {e}")
                        continue
                
                # Альтернативні селектори для категорій
                category_selectors = [
                    f'a[href*="/c{category_id}/"]',
                    f'a[href*="/ua/c{category_id}/"]',
                    f'*[data-category-id="{category_id}"]',
                    f'*[data-id="{category_id}"]'
                ]
                
                for selector in category_selectors:
                    try:
                        elements = soup.select(selector)
                        for element in elements:
                            text = element.get_text(strip=True)
                            if text and len(text) > 2 and len(text) < 100:
                                if not any(skip in text.lower() for skip in ['>', '<', 'function', 'script', 'style']):
                                    if self.debug:
                                        print(f"[parse_category] Знайдено за селектором '{selector}': '{text}'")
                                    return text
                    except Exception as e:
                        if self.debug:
                            print(f"[parse_category] Помилка селектора {selector}: {e}")
                        continue
            
            # Regex пошук як останній варіант
            regex_patterns = [
                rf'<a[^>]+href="[^"]*/?c{category_id}/[^"]*"[^>]*>([^<]+)</a>',
                rf'<a[^>]+href="[^"]*c{category_id}[^"]*"[^>]*>([^<]*?)</a>',
                rf'"name"\s*:\s*"([^"]*)"[^}}]*"categoryId"\s*:\s*"?{category_id}"?',
                rf'"title"\s*:\s*"([^"]*)"[^}}]*"id"\s*:\s*"?{category_id}"?'
            ]
            
            for pattern in regex_patterns:
                try:
                    matches = re.finditer(pattern, html, re.I | re.S)
                    for match in matches:
                        text = re.sub(r'<[^>]+>', '', match.group(1)).strip()
                        text = re.sub(r'\s+', ' ', text)
                        if text and len(text) > 2 and len(text) < 100:
                            if not any(skip in text.lower() for skip in ['function', 'script', 'style', '{', '}']): 
                                if self.debug:
                                    print(f"[parse_category] Знайдено regex: '{text}'")
                                return text
                except Exception as e:
                    if self.debug:
                        print(f"[parse_category] Помилка pattern: {e}")
                    continue
            
            if self.debug:
                print(f"[parse_category] Категорію з ID {category_id} не знайдено")
                
        except Exception as e:
            if self.debug:
                print(f"[parse_category] Загальна помилка: {e}")
        
        return None

    def get_product_meta(self, product_url, add_data, product_id):
        """Покращена функція отримання метаданих товару"""
        title = None
        category_id = None
        category_name = None
        original_url = product_url
        
        # Спочатку намагаємося отримати дані з API відповіді корзини
        if add_data:
            goods_items = add_data.get('purchases', {}).get('goods')
            if goods_items:
                for item in goods_items:
                    goods = item.get('goods', {})
                    if goods.get('id') == product_id:
                        title = goods.get('title') or goods.get('name') or None
                        category_id = goods.get('category_id') or None
                        
                        # Очищаем название от лишних символов
                        if title:
                            title = re.sub(r'\s+', ' ', title).strip()
                        
                        # Оновлюємо URL якщо є кращий варіант
                        api_url = goods.get('href') or goods.get('url')
                        if api_url and api_url.startswith('http'):
                            product_url = api_url
                        break
        
        # Якщо не вдалося отримати з API, пробуємо парсинг HTML
        if not title or not category_id:
            try:
                resp = self.scraper.get(original_url, timeout=15)
                html = resp.text
                
                # Поиск названия товара
                if not title and _HAVE_BS4:
                    soup = BeautifulSoup(html, 'html.parser')
                    
                    title_selectors = [
                        'h1.product__title',
                        'h1[data-testid="product-title"]',
                        '.product-title h1',
                        'h1.rz-product-title',
                        '.product__title',
                        '[data-testid="product-title"]',
                        'h1'
                    ]
                    
                    for selector in title_selectors:
                        try:
                            element = soup.select_one(selector)
                            if element:
                                title = element.get_text(strip=True)
                                if title and len(title) > 5:
                                    title = re.sub(r'\s+', ' ', title).strip()
                                    break
                        except:
                            continue
                
                # Поиск category_id если не найден в API
                if not category_id:
                    # Шукаємо в поточному URL
                    for url_to_check in [product_url, original_url]:
                        match = re.search(r'/c(\d+)/', url_to_check)
                        if match:
                            category_id = int(match.group(1))
                            break
                    
                    # Дополнительный поиск в HTML
                    if not category_id:
                        category_patterns = [
                            r'"category_id"\s*:\s*(\d+)',
                            r'"categoryId"\s*:\s*(\d+)',
                            r'data-category-id="(\d+)"',
                            r'category[_-]?id["\']?\s*[:=]\s*["\']?(\d+)'
                        ]
                        
                        for pattern in category_patterns:
                            match = re.search(pattern, html, re.I)
                            if match:
                                try:
                                    category_id = int(match.group(1))
                                    break
                                except:
                                    continue
                    
            except Exception as e:
                if self.debug:
                    print(f"[get_product_meta] Помилка парсингу HTML: {e}")
        
        # Отримуємо назву категорії більш детально
        if category_id is not None:
            # Спочатку пробуємо отримати з breadcrumbs
            category_name = self.get_category_from_breadcrumbs(product_url, category_id)
            
            # Якщо не вдалося, пробуємо загальний парсинг
            if not category_name:
                category_name = self.parse_category_from_html(product_url, category_id)
            
            # Якщо все ще не вдалося, пробуємо API Rozetka
            if not category_name:
                category_name = self.get_category_from_api(category_id)
                
            # Останній fallback
            if not category_name:
                category_name = f"Категорія #{category_id}"
        
        # Остання перевірка та очистка даних
        if title:
            title = title[:200]
        if category_name:
            category_name = category_name[:100]
            # Видаляємо технічні символи
            category_name = re.sub(r'[<>{}"\']', '', category_name).strip()
            
        if self.debug:
            print(f"[get_product_meta] Результат: title='{title}', category='{category_name}', category_id={category_id}")
            
        return title, category_name

    def get_category_from_breadcrumbs(self, product_url, category_id):
        """Отримання категорії з breadcrumbs з покращеним парсингом"""
        try:
            resp = self.scraper.get(product_url, timeout=15)
            html = resp.text
            
            if self.debug:
                print(f"[breadcrumbs] Шукаємо категорію ID: {category_id}")
            
            if _HAVE_BS4:
                soup = BeautifulSoup(html, 'html.parser')
                
                # Розширені селектори для breadcrumbs
                breadcrumb_selectors = [
                    '.breadcrumbs a',
                    '.rz-breadcrumbs a',
                    '[data-testid="breadcrumbs"] a',
                    '.catalog-heading a',
                    '.breadcrumb a',
                    'nav a',
                    '.rz-catalog-breadcrumbs a',
                    'a[rzrelnofollow].black-link',  # Добавлен селектор для Angular-компонентов
                    'a[href*="/c"]'
                ]
                
                for selector in breadcrumb_selectors:
                    try:
                        links = soup.select(selector)
                        for link in links:
                            href = link.get('href', '')
                            
                            # Перевіряємо чи містить наш category_id
                            if f'/c{category_id}/' in href or f'c{category_id}' in href:
                                # Отримуємо текст, очищаючи від SVG та інших елементів
                                text_content = []
                                
                                # Проходимо по всіх текстових вузлах
                                for text_node in link.find_all(text=True):
                                    text = text_node.strip()
                                    # Пропускаємо порожні рядки та технічні елементи
                                    if text and not text.startswith('icon-') and len(text) > 2:
                                        text_content.append(text)
                                
                                # Об'єднуємо текст
                                full_text = ' '.join(text_content).strip()
                                
                                # Очищаємо від зайвих символів
                                full_text = re.sub(r'\s+', ' ', full_text)
                                full_text = re.sub(r'^[^\w]*|[^\w]*$', '', full_text)  # Видаляємо символи на початку/кінці
                                
                                if full_text and len(full_text) > 2 and len(full_text) < 100:
                                    # Додаткова перевірка що це не технічний текст
                                    if not any(skip in full_text.lower() for skip in [
                                        'svg', 'icon', 'chevron', 'use', 'href', 'assets', 'sprite'
                                    ]):
                                        if self.debug:
                                            print(f"[breadcrumbs] Знайдено категорію: '{full_text}' в {href}")
                                        return full_text
                                
                    except Exception as e:
                        if self.debug:
                            print(f"[breadcrumbs] Помилка селектора {selector}: {e}")
                        continue
            
            # Якщо BeautifulSoup не допоміг, пробуємо regex
            regex_patterns = [
                # Для Angular компонентів з rzrelnofollow
                rf'<a[^>]*rzrelnofollow[^>]*href="[^"]*c{category_id}[^"]*"[^>]*>(?:(?!</a>).)*?([^<>{{}}]+?)(?:(?!</a>).)*?</a>',
                # Звичайні посилання
                rf'<a[^>]*href="[^"]*c{category_id}[^"]*"[^>]*>([^<]+)</a>',
                # Більш складні випадки з вкладеними тегами
                rf'<a[^>]*href="[^"]*c{category_id}[^"]*"[^>]*>(.*?)</a>',
            ]
            
            for pattern in regex_patterns:
                try:
                    matches = re.finditer(pattern, html, re.I | re.S)
                    for match in matches:
                        raw_text = match.group(1)
                        
                        # Видаляємо HTML теги
                        clean_text = re.sub(r'<[^>]+>', '', raw_text)
                        # Видаляємо зайві пробіли
                        clean_text = re.sub(r'\s+', ' ', clean_text).strip()
                        # Видаляємо спеціальні символи на початку/кінці
                        clean_text = re.sub(r'^[^\w\u0400-\u04FF]*|[^\w\u0400-\u04FF]*$', '', clean_text)
                        
                        if clean_text and len(clean_text) > 2 and len(clean_text) < 100:
                            # Перевіряємо що це не технічний текст
                            if not any(skip in clean_text.lower() for skip in [
                                'svg', 'icon', 'chevron', 'use', 'href', 'assets', 'sprite', 'ng-', 'function'
                            ]):
                                if self.debug:
                                    print(f"[breadcrumbs] Знайдено regex: '{clean_text}'")
                                return clean_text
                                
                except Exception as e:
                    if self.debug:
                        print(f"[breadcrumbs] Помилка regex: {e}")
                    continue
            
            if self.debug:
                print(f"[breadcrumbs] Категорію для ID {category_id} не знайдено")
                
        except Exception as e:
            if self.debug:
                print(f"[breadcrumbs] Загальна помилка: {e}")
        
        return None

    def get_category_from_api(self, category_id):
        """Спроба отримати категорію через API Rozetka"""
        try:
            api_url = f"https://common-api.rozetka.com.ua/v2/fat-menu/full?country=UA&lang=ua"
            resp = self.scraper.get(api_url, timeout=10)
            
            if resp.status_code == 200:
                data = resp.json()
                
                def find_category_recursive(items, target_id):
                    for item in items:
                        if item.get('id') == target_id:
                            return item.get('title', item.get('name', ''))
                        
                        children = item.get('children', [])
                        if children:
                            result = find_category_recursive(children, target_id)
                            if result:
                                return result
                    return None
                
                result = find_category_recursive(data.get('data', []), category_id)
                if result:
                    return result
                    
        except Exception as e:
            if self.debug:
                print(f"[get_category_from_api] Помилка: {e}")
        
        return None

    def check_product(self, product_url):
        # ВАЖЛИВО: Очищаємо стан перед перевіркою кожного товару
        self.reset_session_state()
        
        product_id = self.extract_product_id(product_url)
        if not product_id:
            return {"error": "Не вдалося витягти ID", "url": product_url}

        print(f"=== Перевіряємо товар ID {product_id}: {product_url}")
        
        max_stock, add_data = self.binary_search_max_stock(product_id)
        if max_stock is None:
            return {"error": "Не вдалося визначити кількість", "url": product_url, "product_id": product_id}

        title, category_name = self.get_product_meta(product_url, add_data, product_id)
        
        result = {
            "product_id": product_id,
            "url": product_url,
            "title": title or '',
            "category": category_name or '',
            "max_stock": max_stock,
        }
        
        print(f"✅ Результат для товару {product_id}: {max_stock} шт. | {title or 'Без назви'}")
        return result


# Решта коду залишається без змін...
EXCEL_FILENAME = "rozetka_stock_history.xlsx"
EXCEL_FIELDS = ["name", "url", "category", "last_checked", "max_stock"]


def load_existing_excel(path: str):
    """Загружает данные из Excel в список словарей"""
    if not os.path.exists(path):
        return []
    
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True)
        worksheet = workbook.active
        
        # Читаем заголовки из первой строки
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(cell.value)
            else:
                break
        
        if not headers:
            return []
        
        # Читаем данные
        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Пропускаем пустые строки
                continue
            
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value if value is not None else ''
                    
            # Добавляем недостающие поля
            for field in EXCEL_FIELDS:
                if field not in row_dict:
                    row_dict[field] = ''
                    
            data.append(row_dict)
        
        workbook.close()
        return data
        
    except Exception as e:
        print(f"[ПОПЕРЕДЖЕННЯ] Не вдалося завантажити існуючий Excel файл: {e}")
        print("Створюємо новий файл...")
        return []



def save_excel_with_formatting(path: str, data_list):
    """Сохраняет список словарей в Excel с форматированием"""
    if not data_list:
        print("[ПОПЕРЕДЖЕННЯ] Список даних порожній, створюємо файл тільки з заголовками")
        data_list = []
    
    # Группируем данные по товарам и собираем историю
    products_history = {}
    all_dates = set()
    
    for row in data_list:
        url = row.get('url', '')
        if url not in products_history:
            products_history[url] = {
                'name': row.get('name', ''),
                'category': row.get('category', ''),
                'url': url,
                'dates': {}
            }
        
        date = row.get('last_checked', '')
        if date:
            date_only = date.split(' ')[0] if ' ' in date else date
            products_history[url]['dates'][date_only] = row.get('max_stock', 0)
            all_dates.add(date_only)
    
    sorted_dates = sorted(list(all_dates))
    
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Істория залишків"

    # Заголовки
    headers = ["Назва", "URL", "Категорія"] + sorted_dates
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        
        cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Данные
    row_num = 2
    for product_data in products_history.values():
        # Название товара
        cell = ws.cell(row=row_num, column=1, value=product_data['name'])
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # URL
        cell = ws.cell(row=row_num, column=2, value=product_data['url'])
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Категория
        cell = ws.cell(row=row_num, column=3, value=product_data['category'])
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Данные по датам
        for col_idx, date in enumerate(sorted_dates, 4):
            stock_value = product_data['dates'].get(date, '')
            cell = ws.cell(row=row_num, column=col_idx, value=stock_value)
            
            cell.font = Font(name='Arial', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Цветовое кодирование
            if stock_value and stock_value > 0:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif stock_value == 0:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # Чередование строк
            if row_num % 2 == 0:
                if not cell.fill.start_color or cell.fill.start_color.rgb == '00000000':
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        row_num += 1

    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 40  # Название товара
    ws.column_dimensions['B'].width = 60  # URL
    ws.column_dimensions['C'].width = 25  # Категория
    
    # Для столбцов с датами
    for col_idx in range(4, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 12

    # Высота строк
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 25

    # Закрепляем первую строку
    ws.freeze_panes = 'A2'
    
    # Автофильтр
    max_row = len(products_history) + 1
    max_col = len(headers)
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row}"

    try:
        wb.save(path)
    except Exception as e:
        print(f"[ОШИБКА] Не удалось сохранить Excel файл: {e}")
        raise

def upsert_rows(existing_data, new_items):
    """Обновляет список данных новыми элементами"""
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    new_rows = []
    
    for item in new_items:
        if 'error' in item:
            print(f"[ПОПЕРЕДЖЕННЯ] {item.get('url', 'Невідомий URL')}: {item['error']}")
            continue
        new_rows.append({
            'name': item.get('title', ''),
            'url': item.get('url', ''),
            'category': item.get('category', ''),
            'last_checked': now_str,
            'max_stock': item.get('max_stock', 0),
        })
    
    if not existing_data:
        existing_data = []
    
    # Удаляем старые записи для тех же URL
    new_urls = [row['url'] for row in new_rows]
    filtered_existing = [row for row in existing_data if row.get('url', '') not in new_urls]
    
    # Добавляем новые записи
    filtered_existing.extend(new_rows)
    
    return filtered_existing


def read_urls_from_file(fname):
    urls = []
    with open(fname, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            urls.append(line)
    return urls


def get_interactive_urls():
    print("\n" + "="*70)
    print("🛒 ROZETKA STOCK CHECKER - Інтерактивний режим")
    print("="*70)
    print("📝 Введіть URL товарів для перевірки залишків:")
    print("   • Вводьте по одному URL в рядку")
    print("   • Для завершення натисніть Enter на порожньому рядку")
    print("   • Для виходу введіть 'exit' або 'quit'")
    print("-"*70)
    
    urls = []
    counter = 1
    
    while True:
        try:
            url = input(f"🔗 URL №{counter}: ").strip()
            
            if not url:
                if urls:
                    print(f"\n✅ Введено {len(urls)} URL(s). Починаємо перевірку...")
                    break
                else:
                    print("❌ Не введено жодного URL. Спробуйте ще раз.")
                    continue
            
            if url.lower() in ['exit', 'quit', 'вихід']:
                print("👋 Вихід з програми...")
                sys.exit(0)
            
            if url.startswith('http') and 'rozetka.com.ua' in url:
                urls.append(url)
                print(f"   ✓ URL №{counter} додано")
                counter += 1
            else:
                print("   ❌ URL має починатися з http:// або https:// та містити rozetka.com.ua")
                
        except KeyboardInterrupt:
            print("\n\n👋 Програма перервана користувачем")
            sys.exit(0)
    
    return urls

def get_interactive_urls():
    # Проверяем, запущены ли мы в интерактивном терминале
    if not sys.stdin.isatty():
        print("❌ Интерактивный режим недоступен в неинтерактивной среде")
        print("💡 Используйте аргументы командной строки или файл с URL")
        return []
    
    print("\n" + "="*70)
    print("🛒 ROZETKA STOCK CHECKER - Інтерактивний режим")
    print("="*70)
    print("📝 Введіть URL товарів для перевірки залишків:")
    print("   • Вводьте по одному URL в рядку")
    print("   • Для завершення натисніть Enter на порожньому рядку")
    print("   • Для виходу введіть 'exit' або 'quit'")
    print("-"*70)
    
    urls = []
    counter = 1
    
    while True:
        try:
            url = input(f"🔗 URL №{counter}: ").strip()
            
            if not url:
                if urls:
                    print(f"\n✅ Введено {len(urls)} URL(s). Починаємо перевірку...")
                    break
                else:
                    print("❌ Не введено жодного URL. Спробуйте ще раз.")
                    continue
            
            if url.lower() in ['exit', 'quit', 'вихід']:
                print("👋 Вихід з програми...")
                sys.exit(0)
            
            if url.startswith('http') and 'rozetka.com.ua' in url:
                urls.append(url)
                print(f"   ✓ URL №{counter} додано")
                counter += 1
            else:
                print("   ❌ URL має починатися з http:// або https:// та містити rozetka.com.ua")
                
        except KeyboardInterrupt:
            print("\n\n👋 Програма перервана користувачем")
            sys.exit(0)
        except EOFError:
            print("\n❌ Помилка вводу. Завершення роботи.")
            break
    
    return urls

def parse_cli():
    p = argparse.ArgumentParser(description="Rozetka stock checker -> Excel таблиця")
    p.add_argument('urls', nargs='*', help='URL товарів')
    p.add_argument('-f', '--file', help='Файл зі списком URL (по 1 в рядку)')
    p.add_argument('--interactive', action='store_true', help='Інтерактивний режим для вводу URL')
    p.add_argument('--debug', action='store_true', help='Дебаг вивід')
    p.add_argument('--delay', type=float, default=0.7, help='Затримка між запитами під час бінарного пошуку')
    return p.parse_args()


def main():
    print("🚀 Запуск Rozetka Stock Checker...")
    
    args = parse_cli()
    urls = list(args.urls)
    
    if args.file:
        print(f"📄 Завантажуємо URL з файлу: {args.file}")
        urls.extend(read_urls_from_file(args.file))
    
    if not urls:
        print("🔄 Запускається інтерактивний режим...")
        urls = get_interactive_urls()
    
    if not urls:
        print("❌ Не знайдено URL для перевірки!")
        return

    print(f"\n🎯 Знайдено {len(urls)} товарів для перевірки")
    print("⏳ Починаємо перевірку залишків...\n")

    checker = RozetkaStockChecker(debug=args.debug, delay=args.delay)
    results = []
    
    for i, url in enumerate(urls, 1):
        print(f"[{i}/{len(urls)}] Перевіряємо товар...")
        res = checker.check_product(url)
        results.append(res)
        
        if i < len(urls):
            print("⏱️  Пауза між запитами...")
            time.sleep(2)  # Збільшено паузу між товарами

    existing = load_existing_excel(EXCEL_FILENAME)
    merged = upsert_rows(existing, results)
    
    save_excel_with_formatting(EXCEL_FILENAME, merged)

    print("\n" + "="*70)
    print("✅ ГОТОВО! Результати перевірки:")
    print("="*70)
    print(f"📊 Файл збережено: {os.path.abspath(EXCEL_FILENAME)}")
    print(f"📈 Всього записів в таблиці: {len(merged)}")
    print("-"*70)
    

    success_count = 0
    for item in results:
        if 'error' in item:
            print(f"❌ ПОМИЛКА: {item['url']} - {item['error']}")
        else:
            success_count += 1
            print(f"✅ {item['title']}")
            print(f"   📂 Категорія: {item['category']}")
            print(f"   📦 Максимальна кількість: {item['max_stock']}")
            print(f"   🔗 URL: {item['url'][:60]}...")
            print()
    
    print("="*70)
    print(f"🎉 Успішно оброблено: {success_count}/{len(results)} товарів")


if __name__ == '__main__':
    main()
