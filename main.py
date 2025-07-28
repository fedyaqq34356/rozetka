import asyncio
import logging
import os
import re
import sqlite3
from datetime import datetime, time
from typing import List, Dict, Optional
import tempfile
import openpyxl.utils


from dotenv import load_dotenv
load_dotenv()

from aiogram import Bot, Dispatcher, F
from aiogram.types import Message, CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.types import FSInputFile

from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.storage.memory import MemoryStorage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from tg import RozetkaStockChecker, load_existing_excel, save_excel_with_formatting, upsert_rows, EXCEL_FILENAME

# Налаштування логування
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Токен бота (завантажується з .env)
BOT_TOKEN = os.getenv("BOT_TOKEN")


# Визначення станів для FSM
class BotStates(StatesGroup):
    waiting_url = State()
    waiting_time = State()

# Покращений клас для роботи з Rozetka
class ImprovedRozetkaChecker(RozetkaStockChecker):
    def __init__(self, debug=False, delay=2):
        super().__init__(debug, delay)

# Виправлений клас для роботи з базою даних
class DatabaseManager:
    def __init__(self, db_path: str = "rozetka_bot.db"):
        self.db_path = db_path
        self.init_database()

    def init_database(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Таблица продуктов (убираем last_stock и last_check)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                url TEXT UNIQUE NOT NULL,
                name TEXT,
                category TEXT,
                added_date DATE DEFAULT CURRENT_DATE
            )
        """)

        # Таблица истории остатков (основная таблица с данными по датам)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS stock_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER,
                check_date DATE,
                stock_count INTEGER,
                FOREIGN KEY (product_id) REFERENCES products (id),
                UNIQUE(product_id, check_date)
            )
        """)

        # Таблица настроек
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """)

        conn.commit()
        conn.close()

    def add_product(self, url: str, name: str = "", category: str = "") -> bool:
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("""
                INSERT OR REPLACE INTO products (url, name, category, added_date) 
                VALUES (?, ?, ?, CURRENT_DATE)
            """, (url, name, category))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            logger.error(f"Помилка додавання товару: {e}")
            return False

    def get_product_id_by_url(self, url: str) -> Optional[int]:
        """Отримати ID товару по URL"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM products WHERE url = ?", (url,))
        result = cursor.fetchone()
        conn.close()
        return result[0] if result else None

    def update_product_stock(self, product_id: int, stock_count: int):
        """Обновить остатки товара на текущую дату"""
        try:
            logger.info(f"[DB] Начинаем обновление остатков: product_id={product_id}, stock_count={stock_count}")

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Проверяем, существует ли товар
            cursor.execute("SELECT id, name FROM products WHERE id = ?", (product_id,))
            product = cursor.fetchone()
            if not product:
                logger.error(f"[DB] Товар с ID {product_id} не найден в базе данных")
                conn.close()
                return False

            logger.info(f"[DB] Товар найден: ID={product[0]}, Name='{product[1]}'")

            today = datetime.now().strftime('%Y-%m-%d')
            logger.info(f"[DB] Текущая дата: {today}")

            # Проверяем, есть ли уже запись на сегодня
            cursor.execute("SELECT stock_count FROM stock_history WHERE product_id = ? AND check_date = ?",
                           (product_id, today))
            existing = cursor.fetchone()

            if existing:
                logger.info(
                    f"[DB] Найдена существующая запись на {today}: old_stock={existing[0]}, new_stock={stock_count}")
            else:
                logger.info(f"[DB] Новая запись на {today}: stock={stock_count}")

            # Добавляем или обновляем запись в истории
            cursor.execute("""
                INSERT OR REPLACE INTO stock_history (product_id, check_date, stock_count) 
                VALUES (?, ?, ?)
            """, (product_id, today, stock_count))

            affected_rows = cursor.rowcount
            logger.info(f"[DB] Затронуто строк: {affected_rows}")

            conn.commit()
            conn.close()

            logger.info(f"[DB] ✅ УСПЕШНО обновлены остатки для товара {product_id}: {stock_count}")
            return True

        except Exception as e:
            logger.error(f"[DB] ❌ ОШИБКА обновления остатков для товара {product_id}: {e}")
            import traceback
            logger.error(f"[DB] Полный traceback: {traceback.format_exc()}")
            try:
                conn.close()
            except:
                pass
            return False

    def get_products(self) -> List[Dict]:
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT p.id, p.url, p.name, p.category,
                (SELECT stock_count FROM stock_history sh 
                    WHERE sh.product_id = p.id 
                    ORDER BY sh.check_date DESC LIMIT 1) as last_stock,
                (SELECT check_date FROM stock_history sh 
                    WHERE sh.product_id = p.id 
                    ORDER BY sh.check_date DESC LIMIT 1) as last_check
            FROM products p
            ORDER BY p.name
        """)
        products = []
        for row in cursor.fetchall():
            products.append({
                "id": row[0], 
                "url": row[1], 
                "name": row[2] or "Без названия", 
                "category": row[3] or "Без категории",
                "last_stock": row[4] or 0,
                "last_check": row[5] or "Никогда"
            })
        conn.close()
        return products

    def remove_product_by_id(self, product_id: int) -> bool:
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM stock_history WHERE product_id = ?", (product_id,))
            cursor.execute("DELETE FROM products WHERE id = ?", (product_id,))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            logger.error(f"Помилка видалення товару: {e}")
            return False


    def get_products_with_history(self) -> List[Dict]:
        """Получить товары с историей по всем датам"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Получаем все уникальные даты
        cursor.execute("SELECT DISTINCT check_date FROM stock_history ORDER BY check_date")
        all_dates = [row[0] for row in cursor.fetchall()]
        
        # Получаем товары с историей
        cursor.execute("""
            SELECT p.id, p.name, p.url, p.category
            FROM products p
            ORDER BY p.name
        """)
        
        products_data = []
        for product_row in cursor.fetchall():
            product_id, name, url, category = product_row
            
            # Получаем историю остатков для этого товара
            cursor.execute("""
                SELECT check_date, stock_count 
                FROM stock_history 
                WHERE product_id = ? 
                ORDER BY check_date
            """, (product_id,))
            
            history = dict(cursor.fetchall())
            
            product_data = {
                'name': name or 'Без названия',
                'url': url,
                'category': category or 'Без категории',
                'history': history,
                'all_dates': all_dates
            }
            products_data.append(product_data)
        
        conn.close()
        return products_data



    def get_product_by_id(self, product_id: int) -> Optional[Dict]:
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT id, url, name, category FROM products WHERE id = ?", (product_id,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            return {"id": result[0], "url": result[1], "name": result[2], "category": result[3]}
        return None

    def get_schedule_time(self) -> Optional[str]:
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT value FROM settings WHERE key = 'schedule_time'")
        result = cursor.fetchone()
        conn.close()
        return result[0] if result else None

    def set_schedule_time(self, time_str: str):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('schedule_time', ?)", (time_str,))
        conn.commit()
        conn.close()

    def sync_with_excel(self):
        """Синхронізація з Excel файлом"""
        try:
            if os.path.exists(EXCEL_FILENAME):
                data_list = load_existing_excel(EXCEL_FILENAME)
                logger.info(f"Завантажено {len(data_list)} записів з Excel")
                
                for row in data_list:
                    url = row.get('url', '')
                    if url:  # Проверяем, что URL не пустой
                        # Додаємо товар в базу даних
                        self.add_product(
                            url=str(url),
                            name=str(row.get('name', '')),
                            category=str(row.get('category', ''))
                        )
                        
                        # Оновлюємо залишки якщо є дані
                        max_stock = row.get('max_stock')
                        if max_stock is not None and max_stock != '':
                            product_id = self.get_product_id_by_url(str(url))
                            if product_id:
                                try:
                                    stock = int(max_stock)
                                    self.update_product_stock(product_id, stock)
                                except (ValueError, TypeError):
                                    pass  # Пропускаем некорректные значения
                
                logger.info("Синхронізація з Excel завершена")
        except Exception as e:
            logger.error(f"Помилка синхронізації з Excel: {e}")



    def export_to_excel(self):
        """Експорт даних в Excel для main.py"""
        try:
            products = self.get_products()
            excel_data = []
            
            for product in products:
                excel_data.append({
                    'name': product['name'],
                    'url': product['url'],
                    'category': product['category'],
                    'last_checked': product['last_check'],
                    'max_stock': product['last_stock']
                })
            
            if excel_data:
                # Завантажуємо існуючий Excel
                existing_data = load_existing_excel(EXCEL_FILENAME)
                
                # Оновлюємо дані
                updated_data = upsert_rows(existing_data, excel_data)
                
                # Зберігаємо
                save_excel_with_formatting(EXCEL_FILENAME, updated_data)
                logger.info(f"Експортовано {len(excel_data)} товарів в Excel")
                
        except Exception as e:
            logger.error(f"Помилка експорту в Excel: {e}")

class RozetkaTelegramBot:
    def __init__(self):
        self.bot = Bot(token=BOT_TOKEN)
        self.dp = Dispatcher(storage=MemoryStorage())
        self.db = DatabaseManager()
        self.checker = ImprovedRozetkaChecker(debug=True, delay=0.7)  # Enable debug
        self.setup_handlers()
        self.db.sync_with_excel()

    def setup_handlers(self):
        self.dp.message(Command("start"))(self.cmd_start)
        self.dp.message(Command("help"))(self.cmd_help)
        self.dp.message(Command("add"))(self.cmd_add_url)
        self.dp.message(Command("list"))(self.cmd_list_products)
        self.dp.message(Command("remove"))(self.cmd_remove_product)
        self.dp.message(Command("schedule"))(self.cmd_set_schedule)
        self.dp.message(Command("check"))(self.cmd_manual_check)
        self.dp.message(Command("export"))(self.cmd_export_table)
        self.dp.message(Command("sync"))(self.cmd_sync_excel)  # Нова команда
        self.dp.message(F.text)(self.handle_text_messages)
        self.dp.callback_query()(self.handle_callback_query)

    async def cmd_sync_excel(self, message: Message):
        """Синхронізація з Excel файлом"""
        await message.reply("🔄 Синхронізую з Excel файлом...")
        
        try:
            self.db.sync_with_excel()
            products_count = len(self.db.get_products())
            await message.reply(f"✅ Синхронізація завершена!\n📊 Всього товарів: {products_count}")
        except Exception as e:
            await message.reply(f"❌ Помилка синхронізації: {str(e)}")

    async def cmd_start(self, message: Message):
        await message.reply(
            "🛒 <b>Бот перевірки залишків Rozetka</b>\n\n"
            "📋 Доступні команди:\n"
            "/add - додати товар\n"
            "/list - список товарів\n"
            "/remove - видалити товар\n"
            "/schedule - налаштувати розклад\n"
            "/check - ручна перевірка\n"
            "/export - експорт таблиці\n"
            "/sync - синхронізація з Excel\n"
            "/help - допомога",
            parse_mode="HTML"
        )

    async def cmd_help(self, message: Message):
        await message.reply(
            "🔧 <b>Інструкція:</b>\n\n"
            "1. Додайте товари командою /add\n"
            "2. Встановіть час перевірки /schedule\n"
            "3. Бот щодня перевірятиме залишки\n"
            "4. Експортуйте дані /export\n"
            "5. Синхронізуйте з Excel /sync\n\n"
            "⚠️ Формат часу: ГГ:ХХ (наприклад, 09:30)",
            parse_mode="HTML"
        )

    async def cmd_add_url(self, message: Message, state: FSMContext):
        await state.set_state(BotStates.waiting_url)
        await message.reply("🔗 Надішліть посилання на товар Rozetka:")

    async def cmd_list_products(self, message: Message):
        products = self.db.get_products()
        if not products:
            await message.reply("📦 Список товарів порожній")
            return
        
        text = "📋 <b>Список товарів:</b>\n\n"
        for i, product in enumerate(products, 1):
            name = product['name']
            category = product['category']
            stock = product['last_stock']
            last_check = product['last_check']
            
            text += f"{i}. <b>{name}</b>\n"
            text += f"   📂 {category}\n"
            text += f"   📊 Залишки: {stock}\n"
            text += f"   🕐 Остання перевірка: {last_check}\n"
            text += f"   🔗 {product['url'][:50]}...\n\n"
        
        await message.reply(text, parse_mode="HTML")

    async def cmd_remove_product(self, message: Message):
        products = self.db.get_products()
        if not products:
            await message.reply("📦 Немає товарів для видалення")
            return
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=f"🗑 {p['name'][:30]}...", 
                                 callback_data=f"remove_{p['id']}")] for p in products
        ])
        await message.reply("Оберіть товар для видалення:", reply_markup=keyboard)

    async def cmd_set_schedule(self, message: Message, state: FSMContext):
        await state.set_state(BotStates.waiting_time)
        current_time = self.db.get_schedule_time()
        text = "🕐 Введіть час щоденної перевірки (формат ГГ:ХХ):"
        if current_time:
            text += f"\n\n⏰ Поточний час: {current_time}"
        await message.reply(text)

    async def cmd_manual_check(self, message: Message):
        await message.reply("🔍 Запускаю ручну перевірку залишків...")
        results = await self.check_products_without_saving()
        
        if results:
            report = "✅ <b>Ручна перевірка завершена!</b>\n\n📊 <b>Результати:</b>\n\n"
            for result in results:
                report += f"📦 <b>{result['name']}</b>\n"
                if result['success']:
                    report += f"   📂 Категорія: {result.get('category', 'Невідома')}\n"
                    report += f"   📈 Залишки: {result['stock']}\n"
                else:
                    report += f"   ❌ Помилка: {result['error']}\n"
                report += "\n"
            
            report += "ℹ️ <i>Дані НЕ збережено в історію. Збереження тільки при автоматичній перевірці.</i>"
            
            if len(report) > 4000:
                await message.reply("✅ Перевірка завершена! Результати надсилаються частинами...")
                chunks = [report[i:i+4000] for i in range(0, len(report), 4000)]
                for chunk in chunks:
                    await message.reply(chunk, parse_mode="HTML")
            else:
                await message.reply(report, parse_mode="HTML")
        else:
            await message.reply("✅ Перевірка завершена, але товарів для перевірки немає")

    async def cmd_export_table(self, message: Message):
        await message.reply("📊 Генерую Excel таблицю...")
        
        try:
            # Спочатку експортуємо в основний Excel файл
            self.db.export_to_excel()
            
            products = self.db.get_products()
            if not products:
                await message.reply("❌ Немає товарів для експорту")
                return
            
            excel_path = await self.generate_excel()
            
            if not os.path.exists(excel_path):
                await message.reply("❌ Помилка створення файлу")
                return
            
            file_size = os.path.getsize(excel_path)
            if file_size == 0:
                await message.reply("❌ Створений файл порожній")
                os.remove(excel_path)
                return
            
            await message.reply_document(
                document=FSInputFile(excel_path, filename="rozetka_stock_history.xlsx"),
                caption=f"📋 Таблиця залишків Rozetka\n📊 Товарів: {len(products)}\n📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            )
            
            os.remove(excel_path)
            logger.info(f"Експорт виконано для {len(products)} товарів")
            
        except Exception as e:
            logger.error(f"Помилка експорту: {e}")
            await message.reply(f"❌ Помилка створення таблиці: {str(e)}")


    async def check_products_without_saving(self) -> List[Dict]:
        """Перевірка товарів БЕЗ збереження в базу даних (для ручної перевірки)"""
        products = self.db.get_products()
        results = []
        
        for i, product in enumerate(products, 1):
            try:
                logger.info(f"Ручна перевірка товару {i}/{len(products)}: {product['name']}")
                
                result = self.checker.check_product(product['url'])
                if 'error' not in result:
                    stock_count = result.get('max_stock', 0)
                    # ИСПРАВЛЕНИЕ: используем данные из result вместо product
                    product_name = result.get('title', product['name'])
                    category_name = result.get('category', 'Без категории')
                    
                    results.append({
                        'name': product_name or 'Без назви',
                        'category': category_name or 'Без категории', # Добавляем категорию
                        'success': True,
                        'stock': stock_count
                    })
                    
                    logger.info(f"Ручна перевірка - Успіх: {product_name}, категория: {category_name}, залишки: {stock_count}")
                else:
                    results.append({
                        'name': product['name'],
                        'category': 'Помилка',
                        'success': False,
                        'error': result['error']
                    })
                    logger.error(f"Ручна перевірка - Помилка для товару {product['url']}: {result['error']}")
                    
            except Exception as e:
                logger.error(f"Критична помилка ручної перевірки товару {product['url']}: {e}")
                results.append({
                    'name': product['name'],
                    'category': 'Критична помилка',
                    'success': False,
                    'error': str(e)
                })
            
            # Пауза між товарами
            if i < len(products):
                await asyncio.sleep(2)
        
        return results

    async def generate_excel(self) -> str:
        temp_dir = tempfile.gettempdir()
        filename = f"rozetka_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(temp_dir, filename)
        
        try:
            products_data = self.db.get_products_with_history()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Історія залишків"
            
            # Определяем все даты
            all_dates = set()
            for product in products_data:
                all_dates.update(product['all_dates'])
            
            sorted_dates = sorted(list(all_dates))
            
            # Создаем заголовки с колонками изменений
            headers = ["Товар", "URL", "Категорія"]
            for date in sorted_dates:
                headers.extend([f"{date}\nкількість", f"{date}\nзміни"])
            
            # Заполняем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Заполняем данные
            for row_idx, product in enumerate(products_data, 2):
                # Основная информация о товаре
                ws.cell(row=row_idx, column=1, value=product['name'])
                ws.cell(row=row_idx, column=2, value=product['url'])
                ws.cell(row=row_idx, column=3, value=product['category'])
                
                # Заполняем данные по датам
                previous_stock = None
                col_idx = 4
                
                for date in sorted_dates:
                    current_stock = product['history'].get(date, '')
                    
                    # Колонка количества
                    stock_cell = ws.cell(row=row_idx, column=col_idx, value=current_stock)
                    stock_cell.alignment = Alignment(horizontal="center", vertical="center")
                    stock_cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Цветовое кодирование для количества
                    if current_stock and current_stock > 0:
                        stock_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif current_stock == 0:
                        stock_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    # Колонка изменений
                    change_cell = ws.cell(row=row_idx, column=col_idx + 1)
                    change_cell.alignment = Alignment(horizontal="center", vertical="center")
                    change_cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Вычисляем изменения
                    if previous_stock is not None and current_stock != '' and previous_stock != '':
                        try:
                            change = int(current_stock) - int(previous_stock)
                            if change != 0:
                                change_cell.value = change
                                # Цветовое кодирование для изменений
                                if change > 0:
                                    change_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                    change_cell.font = Font(color="006100", bold=True)
                                else:
                                    change_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                    change_cell.font = Font(color="9C0006", bold=True)
                        except (ValueError, TypeError):
                            pass
                    
                    # Обновляем previous_stock для следующей итерации
                    if current_stock != '':
                        previous_stock = current_stock
                    
                    col_idx += 2
            
            # Настройка ширины столбцов
            ws.column_dimensions['A'].width = 40  # Товар
            ws.column_dimensions['B'].width = 60  # URL
            ws.column_dimensions['C'].width = 25  # Категория
            
            # Для колонок с датами и изменениями
            col_idx = 4
            for _ in sorted_dates:
                col_letter_qty = openpyxl.utils.get_column_letter(col_idx)
                col_letter_change = openpyxl.utils.get_column_letter(col_idx + 1)
                ws.column_dimensions[col_letter_qty].width = 12    # Количество
                ws.column_dimensions[col_letter_change].width = 10  # Изменения
                col_idx += 2
            
            # Высота строк
            ws.row_dimensions[1].height = 30  # Заголовок
            for row in range(2, len(products_data) + 2):
                ws.row_dimensions[row].height = 25
            
            # Закрепляем первые строки и столбцы
            ws.freeze_panes = 'D2'
            
            # Автофильтр
            max_row = len(products_data) + 1
            max_col = len(headers)
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row}"
            
            wb.save(filepath)
            logger.info(f"Excel файл створено: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"Помилка створення Excel: {e}")
            # Создаем простой файл с ошибкой
            wb = Workbook()
            ws = wb.active
            ws.title = "Помилка"
            ws.cell(row=1, column=1, value=f"Помилка створення файлу: {str(e)}")
            wb.save(filepath)
            return filepath
            


    async def handle_callback_query(self, callback_query: CallbackQuery):
        if callback_query.data.startswith("remove_"):
            product_id = int(callback_query.data.split("_")[1])
            product = self.db.get_product_by_id(product_id)
            
            if product:
                success = self.db.remove_product_by_id(product_id)
                if success:
                    # Оновлюємо Excel після видалення
                    self.db.export_to_excel()
                    
                    await callback_query.message.edit_text(
                        f"✅ Товар успішно видалено!\n\n"
                        f"📦 <b>{product['name'] or 'Без назви'}</b>\n"
                        f"🔗 {product['url'][:50]}...",
                        parse_mode="HTML"
                    )
                else:
                    await callback_query.message.edit_text("❌ Помилка видалення товару")
            else:
                await callback_query.message.edit_text("❌ Товар не знайдено")
        
        await callback_query.answer()

    async def handle_text_messages(self, message: Message, state: FSMContext):
        current_state = await state.get_state()
        
        if current_state == BotStates.waiting_url:
            await self.process_url(message, state)
        elif current_state == BotStates.waiting_time:
            await self.process_schedule_time(message, state)

    async def process_url(self, message: Message, state: FSMContext):
        url = message.text.strip()
        
        if not re.match(r'https?://.*rozetka\.com\.ua.*', url):
            await message.reply("❌ Невірне посилання. Потрібно посилання на rozetka.com.ua")
            return
        
        processing_msg = await message.reply("⏳ Обробляю товар...")
        
        try:
            result = self.checker.check_product(url)
            
            if 'error' in result:
                await processing_msg.edit_text(f"❌ Помилка: {result['error']}")
                return
            
            # Додаємо товар в базу даних БЕЗ оновлення залишків
            success = self.db.add_product(
                url=result['url'],
                name=result.get('title', 'Без назви'),
                category=result.get('category', 'Невідома')
            )
            
            if success:
                stock = result.get('max_stock', 0)
                category = result.get('category', 'Невідома')
                
                success_text = (
                    f"✅ Товар додано!\n\n"
                    f"📦 <b>{result.get('title', 'Без назви')}</b>\n"
                    f"📂 Категорія: {category}\n"
                    f"📊 Поточні залишки: {stock}\n"
                    f"🔗 URL: {result['url'][:50]}...\n\n"
                    f"ℹ️ Залишки будуть збережені тільки при автоматичній перевірці"
                )
                
                await processing_msg.edit_text(success_text, parse_mode="HTML")
                logger.info(f"Товар додано: {result.get('title', 'Без назви')}, категорія: {category}, URL: {result['url']}")
            else:
                await processing_msg.edit_text("❌ Помилка збереження товару")
                
        except Exception as e:
            logger.error(f"Помилка обробки URL {url}: {e}")
            await processing_msg.edit_text(f"❌ Помилка обробки: {str(e)}")
        
        await state.clear()

    async def check_all_products(self, manual=False) -> List[Dict]:
        products = self.db.get_products()
        results = []

        logger.info(f"=== НАЧАЛО АВТОМАТИЧЕСКОЙ ПРОВЕРКИ ===")
        logger.info(f"Режим manual: {manual}")
        logger.info(f"Всего товаров для проверки: {len(products)}")

        for i, product in enumerate(products, 1):
            try:
                logger.info(f">>> Товар {i}/{len(products)}: {product['name']} (ID: {product['id']})")

                result = self.checker.check_product(product['url'])
                if 'error' not in result:
                    # Оновлюємо інформацію про товар
                    updated_name = result.get('title', product['name'])
                    updated_category = result.get('category', product['category'])
                    stock_count = result.get('max_stock', 0)

                    logger.info(
                        f"    Получен результат: name='{updated_name}', category='{updated_category}', stock={stock_count}")

                    if updated_name != product['name'] or updated_category != product['category']:
                        logger.info(f"    Обновляем информацию о товаре...")
                        self.db.add_product(
                            product['url'],
                            updated_name,
                            updated_category
                        )

                    # Оновлюємо залишки тільки для автоматичних перевірок
                    if not manual:
                        logger.info(f"    СОХРАНЯЕМ ОСТАТКИ для товара ID {product['id']}: {stock_count}")
                        try:
                            success = self.db.update_product_stock(product['id'], stock_count)
                            if success:
                                logger.info(f"    ✅ УСПЕШНО сохранены остатки для товара {product['id']}")
                            else:
                                logger.error(f"    ❌ ОШИБКА сохранения остатков для товара {product['id']}")
                        except Exception as stock_error:
                            logger.error(
                                f"    ❌ ИСКЛЮЧЕНИЕ при сохранении остатков товара {product['id']}: {stock_error}")
                    else:
                        logger.info(f"    Пропускаем сохранение остатков (manual=True)")

                    results.append({
                        'name': updated_name or 'Без назви',
                        'success': True,
                        'stock': stock_count
                    })

                    logger.info(f"    ✅ Товар {i} обработан успешно")
                else:
                    logger.error(f"    ❌ Ошибка проверки товара {i}: {result['error']}")
                    results.append({
                        'name': product['name'],
                        'success': False,
                        'error': result['error']
                    })

            except Exception as e:
                logger.error(f"❌ КРИТИЧЕСКАЯ ОШИБКА для товара {i} ({product.get('name', 'Unknown')}): {e}")
                logger.error(f"   URL: {product.get('url', 'Unknown')}")
                import traceback
                logger.error(f"   Traceback: {traceback.format_exc()}")

                results.append({
                    'name': product.get('name', 'Unknown'),
                    'success': False,
                    'error': str(e)
                })

                # НЕ ПРЕРЫВАЕМ цикл, продолжаем со следующим товаром

            # Пауза між товарами
            if i < len(products):
                logger.info(f"    Пауза перед следующим товаром...")
                await asyncio.sleep(2)

        logger.info(f"=== КОНЕЦ АВТОМАТИЧЕСКОЙ ПРОВЕРКИ ===")
        logger.info(f"Обработано товаров: {len(results)}")
        success_count = sum(1 for r in results if r.get('success', False))
        logger.info(f"Успешно: {success_count}, Ошибок: {len(results) - success_count}")

        return results


    async def process_schedule_time(self, message: Message, state: FSMContext):
        time_text = message.text.strip()
        
        # Проверяем формат времени
        if not re.match(r'^\d{1,2}:\d{2}$', time_text):
            await message.reply("❌ Неправильний формат часу. Використовуйте ГГ:ХХ (наприклад, 09:30)")
            return
        
        try:
            # Проверяем валидность времени
            time.fromisoformat(time_text + ":00")
            
            self.db.set_schedule_time(time_text)
            await message.reply(f"✅ Час щоденної перевірки встановлено: {time_text}")
            
        except ValueError:
            await message.reply("❌ Неправильний час. Використовуйте формат ГГ:ХХ")
        
        await state.clear()


    async def schedule_checker(self):
        """Покращений планувальник з більш точною перевіркою часу"""
        last_check_date = None
        
        while True:
            try:
                schedule_time = self.db.get_schedule_time()
                if schedule_time:
                    now = datetime.now()
                    current_date = now.date()
                    current_time = now.time()
                    
                    # Парсимо час з бази
                    try:
                        target_hour, target_minute = map(int, schedule_time.split(':'))
                        target_time = time(target_hour, target_minute)
                    except ValueError:
                        logger.error(f"Неправильний формат часу в базі: {schedule_time}")
                        await asyncio.sleep(60)
                        continue
                    
                    # Перевіряємо чи потрібно запускати перевірку
                    should_run = (
                        current_date != last_check_date and  # Не запускали сьогодні
                        current_time.hour == target_time.hour and 
                        current_time.minute == target_time.minute
                    )
                    
                    if should_run:
                        logger.info(f"🕐 Запуск планової автоматичної перевірки о {schedule_time}")
                        
                        try:
                            # Запускаємо перевірку всіх товарів
                            results = await self.check_all_products(manual=False)
                            
                            # Експортуємо в Excel після автоматичної перевірки
                            self.db.export_to_excel()
                            
                            # Оновлюємо дату останньої перевірки
                            last_check_date = current_date
                            
                            success_count = sum(1 for r in results if r.get('success', False))
                            logger.info(f"✅ Автоматична перевірка завершена: {success_count}/{len(results)} товарів")
                            
                        except Exception as e:
                            logger.error(f"Помилка автоматичної перевірки: {e}")
                        
                        # Чекаємо 2 хвилини щоб не запускати повторно
                        await asyncio.sleep(120)
                    else:
                        # Звичайна пауза
                        await asyncio.sleep(30)
                else:
                    # Якщо час не встановлено, чекаємо довше
                    await asyncio.sleep(300)
                    
            except Exception as e:
                logger.error(f"Критична помилка планувальника: {e}")
                await asyncio.sleep(300)

    async def start_bot(self):
        logger.info("Запуск Telegram бота")
        asyncio.create_task(self.schedule_checker())
        await self.dp.start_polling(self.bot)

async def main():
    bot_instance = RozetkaTelegramBot()
    await bot_instance.start_bot()

if __name__ == "__main__":
    asyncio.run(main())
